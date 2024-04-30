import openpyxl
import pyautogui
import time

portada = """
 _________________________________________
|              ____                  _____|
|   ________  / / /__  ____  ____   /_   /|
|  / ___/ _ \/ / / _ \/ __ \/ __ \ /_   / |
| / /  /  __/ / /  __/ / / / /_/ //____/  |
|/_/   \___/_/_/\___/_/ /_/\____/         |
|                                         |
|  by: arm-code (GPL License)             |
|_________________________________________|
"""


print(portada)



# FILE 
excel_name = input('Ingrese el nombre del archivo de excel: ')
excel = excel_name + '.xlsx'
workbook = openpyxl.load_workbook(excel, data_only=True)

print("COLOQUE EL CURSOR EN LA VENTANA DEL SIOSAD.")
print('(EN EL ESPACIO DE LA MATRICULA)')
time.sleep(1)

print('\nLEYENDO LOS DATOS DEL EXCEL...')
time.sleep(1)

# Iterar sobre cada hoja del libro
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    num_cols = sheet.max_column
    num_rows = sheet.max_row
    print('***************************************************')
    print('\nLOS SIGUIENTES DATOS HAN SIDO LEIDOS DEL ARCHIVO:')
    print(f'\n\n\nEXAMENES DE LA SEDE #: {sheet_name}')

    estudiantes = input('\n\n\nIngrese la cantidad de estudiantes: ')
    n_estudiantes=int(estudiantes) + 11   

    # Definir el rango de columnas desde B hasta M
    columnas_a_leer = list(sheet.iter_cols(min_col=2, max_col=13, min_row=11, max_row=n_estudiantes))
    datos_estudiante = list(sheet.iter_cols(min_col=14, max_col=16, min_row=11, max_row=n_estudiantes))
    columnas_materias = list(sheet.iter_cols(min_col=17, max_col=20, min_row=11, max_row=n_estudiantes))  
      

    # Iterar sobre cada fila de la hoja
    for row_index in range(2, n_estudiantes-1):  # Solo filas 1 y 2
        matricula = []
        materias = []
        nombres = []
        try:
            # Iterar sobre cada celda en el rango de columnas definido
            for col_index in range(len(columnas_a_leer)):
                cell_value = columnas_a_leer[col_index][row_index - 1].value
                matricula.append(str(cell_value)) 

            for col_index in range(len(datos_estudiante)):
                cell_value = datos_estudiante[col_index][row_index - 1].value
                nombres.append(str(cell_value))           
            
            for col_index in range(len(columnas_materias)):
                cell_value = columnas_materias[col_index][row_index-1].value
                if cell_value != 'None':
                    materias.append(str(cell_value))
        except Exception as err:
            print('\n\n\n\nSIGUIENTE SEDE...')
            break;
        # Imprimir los datos de la fila actual
        print('\n\n\n\n***************************************************')
        print('#:                   ', row_index - 1)
        print('Nombre:              ', ' '.join(map(str, nombres)))
        print('Matricula:           ', ''.join(map(str, matricula)))       
        print('Materias solicitadas:', ' | '.join(map(str, materias)))
        print('Num. sede:           ', sheet_name)
        print('***************************************************')
        print('\n\nINGRESANDO LOS DATOS EN EL SIOSAD...')
        time.sleep(1)
        # se deben verificar las coordenadas del click, de lo contrario se va ir a otro lado la captura
        pyautogui.click(x=150, y=150)              

        # RESPUESTAS
        try:
            for i in range(1,13):            
                pyautogui.write(matricula[i-1])
        except Exception as err:
            print('Ocurrio un error al capturar la matricula', err)            
            break

        # enter para ingresar matricula
        pyautogui.press('enter')    
        # enter para llegar a la cantidad pagada
        pyautogui.press('enter')
        pyautogui.write('372')  #precio de 4 examenes
        pyautogui.press('enter')
        pyautogui.write('S')        
        pyautogui.write(str(sheet_name))
        pyautogui.press('enter')        
        pyautogui.press('enter')
        try:
            for i in range(len(materias)):            
                if materias[i] == 'None':
                    break
                pyautogui.write(materias[i])            
                pyautogui.press('enter')
        except Exception as err:
            print('Ocurrio un error al capturar la materia', err)
            
        
        
        input('\nREVISE CUIDADOSAMENTE LA CAPTURA \nPRESIONE ENTER PARA CONTINUAR...\n>')
        print('enter')
        print('\n\nCOLOQUE EL CURSOR EN EL SIOSAD')
        time.sleep(1)

        # GUARDAR solictud
        pyautogui.press('f2')
        pyautogui.press('enter')
        pyautogui.press('enter')        
        print('\nCAPTURA EXITOSA!\n')             
        
        input('\t| Para detener el programa: \n\t| [PRESIONE CTRL + C]\n\t| Para solicitar la sig. Materia: \n\t| [PRESIONE ENTER]\n\t>')
        
        print('enter')
        time.sleep(2)
        pyautogui.press('f9')        
    workbook.close()